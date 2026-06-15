"""
Duypt2 Task Manager — Import/Export Service.

Handles CSV and Excel import/export for task data.
"""
import csv
import os
from datetime import datetime

from models.task_model import TaskModel
from models.category_model import CategoryModel


class ImportExportService:
    """Handles importing tasks from and exporting tasks to CSV/Excel files."""

    # Expected column headers in import files
    IMPORT_COLUMNS = [
        "Tên công việc", "Mô tả", "Ngày bắt đầu", "Deadline (ngày)",
        "Deadline (giờ)", "Ưu tiên", "Nhóm", "Người phụ trách"
    ]

    PRIORITY_MAP = {
        "cao": "high", "high": "high", "h": "high",
        "trung bình": "medium", "tb": "medium", "medium": "medium", "m": "medium",
        "thấp": "low", "low": "low", "l": "low",
    }

    def __init__(self):
        self.task_model = TaskModel()
        self.category_model = CategoryModel()

    def generate_template_csv(self, filepath: str):
        """Generate a CSV template file with headers and example data."""
        with open(filepath, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.writer(f)
            writer.writerow(self.IMPORT_COLUMNS)
            writer.writerow([
                "Ví dụ: Báo cáo tuần",
                "Báo cáo tiến độ dự án ABC",
                "2025-01-15",
                "2025-01-20",
                "17:00",
                "Cao",
                "Công việc",
                "Nguyễn Văn A"
            ])

    def import_csv(self, filepath: str, user_id: int = 1) -> dict:
        """
        Import tasks from a CSV file.

        Returns:
            dict with keys: success (int), errors (list of str)
        """
        results = {"success": 0, "errors": []}

        if not os.path.exists(filepath):
            results["errors"].append(f"File không tồn tại: {filepath}")
            return results

        categories = {c["name"]: c["id"] for c in self.category_model.get_all()}

        with open(filepath, "r", encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)
            for row_num, row in enumerate(reader, start=2):
                try:
                    title = row.get("Tên công việc", "").strip()
                    if not title:
                        results["errors"].append(f"Dòng {row_num}: Thiếu tên công việc")
                        continue

                    description = row.get("Mô tả", "").strip()
                    start_date = self._parse_date(row.get("Ngày bắt đầu", ""))
                    deadline_date = self._parse_date(row.get("Deadline (ngày)", ""))
                    deadline_time = row.get("Deadline (giờ)", "23:59").strip() or "23:59"

                    if not deadline_date:
                        results["errors"].append(f"Dòng {row_num}: Deadline không hợp lệ")
                        continue

                    priority_raw = row.get("Ưu tiên", "medium").strip().lower()
                    priority = self.PRIORITY_MAP.get(priority_raw, "medium")

                    category_name = row.get("Nhóm", "").strip()
                    category_id = categories.get(category_name)
                    if category_name and not category_id:
                        # Auto-create new category
                        new_cat = self.category_model.create(category_name)
                        if new_cat:
                            category_id = new_cat["id"]
                            categories[category_name] = category_id

                    assignee = row.get("Người phụ trách", "").strip()

                    self.task_model.create_task(
                        title=title,
                        description=description,
                        start_date=start_date,
                        deadline_date=deadline_date,
                        deadline_time=deadline_time,
                        priority=priority,
                        category_id=category_id,
                        assignee=assignee,
                        user_id=user_id,
                    )
                    results["success"] += 1

                except Exception as e:
                    results["errors"].append(f"Dòng {row_num}: {str(e)}")

        return results

    def import_excel(self, filepath: str, user_id: int = 1) -> dict:
        """Import tasks from an Excel (.xlsx) file."""
        results = {"success": 0, "errors": []}

        try:
            import openpyxl
        except ImportError:
            results["errors"].append("Thiếu thư viện openpyxl. Hãy chạy: pip install openpyxl")
            return results

        if not os.path.exists(filepath):
            results["errors"].append(f"File không tồn tại: {filepath}")
            return results

        categories = {c["name"]: c["id"] for c in self.category_model.get_all()}

        try:
            wb = openpyxl.load_workbook(filepath)
            ws = wb.active
            headers = [cell.value for cell in ws[1]]

            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    row_dict = dict(zip(headers, row))
                    title = str(row_dict.get("Tên công việc", "")).strip()
                    if not title:
                        continue

                    description = str(row_dict.get("Mô tả", "")).strip()
                    start_date = self._parse_date(row_dict.get("Ngày bắt đầu"))
                    deadline_date = self._parse_date(row_dict.get("Deadline (ngày)"))
                    deadline_time = str(row_dict.get("Deadline (giờ)", "23:59")).strip() or "23:59"

                    if not deadline_date:
                        results["errors"].append(f"Dòng {row_num}: Deadline không hợp lệ")
                        continue

                    priority_raw = str(row_dict.get("Ưu tiên", "medium")).strip().lower()
                    priority = self.PRIORITY_MAP.get(priority_raw, "medium")

                    category_name = str(row_dict.get("Nhóm", "")).strip()
                    category_id = categories.get(category_name)
                    if category_name and not category_id:
                        new_cat = self.category_model.create(category_name)
                        if new_cat:
                            category_id = new_cat["id"]
                            categories[category_name] = category_id

                    assignee = str(row_dict.get("Người phụ trách", "")).strip()

                    self.task_model.create_task(
                        title=title,
                        description=description,
                        start_date=start_date,
                        deadline_date=deadline_date,
                        deadline_time=deadline_time,
                        priority=priority,
                        category_id=category_id,
                        assignee=assignee,
                        user_id=user_id,
                    )
                    results["success"] += 1

                except Exception as e:
                    results["errors"].append(f"Dòng {row_num}: {str(e)}")

        except Exception as e:
            results["errors"].append(f"Lỗi đọc file Excel: {str(e)}")

        return results

    def export_csv(self, filepath: str, tasks: list[dict] = None):
        """Export tasks to a CSV file."""
        if tasks is None:
            tasks = self.task_model.get_all_tasks()

        with open(filepath, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.writer(f)
            writer.writerow(self.IMPORT_COLUMNS + ["Trạng thái", "Tiến độ (%)"])
            for task in tasks:
                writer.writerow([
                    task.get("title", ""),
                    task.get("description", ""),
                    task.get("start_date", ""),
                    task.get("deadline_date", ""),
                    task.get("deadline_time", ""),
                    task.get("priority", ""),
                    task.get("category_name", ""),
                    task.get("assignee", ""),
                    task.get("status", ""),
                    task.get("progress", 0),
                ])

    def _parse_date(self, value) -> str | None:
        """Try to parse a date from various formats into YYYY-MM-DD."""
        if value is None:
            return None

        if isinstance(value, datetime):
            return value.strftime("%Y-%m-%d")

        value = str(value).strip()
        if not value:
            return None

        formats = ["%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d"]
        for fmt in formats:
            try:
                return datetime.strptime(value, fmt).strftime("%Y-%m-%d")
            except ValueError:
                continue
        return None
